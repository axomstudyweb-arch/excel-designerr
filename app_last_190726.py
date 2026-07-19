import os
import uuid
import logging
import io
import time
import shutil
import threading
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

SESSION_DELTAS: Dict[str, Dict[str, Any]] = {}

def run_storage_cleaner():
    while True:
        try:
            now = time.time()
            for item in UPLOAD_DIR.iterdir():
                if item.is_file() and (now - item.stat().st_mtime) > 14400:
                    item.unlink()
                    wb_id = item.stem
                    if wb_id in SESSION_DELTAS:
                        del SESSION_DELTAS[wb_id]
        except Exception as e:
            logger.error(f"Cleaner thread error: {e}")
        time.sleep(1800)

threading.Thread(target=run_storage_cleaner, daemon=True).start()

class HighPerformanceProcessor:
    def parse_csv_with_encoding_fallback(self, file_path: Path, nrows: int = None, skiprows: int = 0) -> pd.DataFrame:
        encodings = ['utf-8', 'utf-8-sig', 'cp1252', 'latin1']
        for enc in encodings:
            try:
                return pd.read_csv(file_path, encoding=enc, header=None, nrows=nrows, skiprows=skiprows)
            except Exception:
                continue
        raise ValueError("Unsupported matrix structure.")

    def find_header_offset(self, file_path: Path, suffix: str) -> int:
        keywords = {'si', 'rc', 'id', 'date', 'name', 'serial', 'no', 'code', 'sl no', 'slno', 'beneficiary', 'farmer'}
        try:
            if suffix == '.csv':
                df = self.parse_csv_with_encoding_fallback(file_path, nrows=30)
            else:
                df = pd.read_excel(file_path, header=None, nrows=30, engine='openpyxl')
            
            for idx, row in df.iterrows():
                row_vals = [str(x).strip().lower() for x in row.values if pd.notna(x)]
                if any(k in row_vals for k in keywords):
                    return idx
        except Exception:
            pass
        return 0

    def get_dimensions(self, file_path: Path, suffix: str):
        try:
            if suffix == '.csv':
                df = self.parse_csv_with_encoding_fallback(file_path, nrows=5)
                cols = len(df.columns)
                total_rows = sum(1 for _ in open(file_path, 'r', encoding='utf-8', errors='ignore'))
            else:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                total_rows = ws.max_row or 100
                cols = ws.max_column or 10
                wb.close()
            return {"rows": max(1, total_rows), "cols": cols}
        except Exception:
            return {"rows": 100, "cols": 10}

processor = HighPerformanceProcessor()

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".xlsx", ".xls", ".csv"}:
        raise HTTPException(status_code=400, detail="Invalid extension format.")
        
    wb_id = str(uuid.uuid4())
    temp_path = UPLOAD_DIR / f"{wb_id}{suffix}"
    
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    offset = processor.find_header_offset(temp_path, suffix)
    dims = processor.get_dimensions(temp_path, suffix)
    
    SESSION_DELTAS[wb_id] = {
        "data": {}, 
        "headers": {},
        "styles": {}, 
        "added_rows_count": 0, 
        "added_cols_count": 0,
        "base_offset": offset,
        "title": ""
    }

    try:
        if suffix == '.csv':
            raw_h = processor.parse_csv_with_encoding_fallback(temp_path, nrows=1, skiprows=offset).values.tolist()[0]
        else:
            df_h = pd.read_excel(temp_path, engine='openpyxl', skiprows=offset, nrows=1, header=None)
            raw_h = df_h.values.tolist()[0]
        headers = [str(x) if pd.notna(x) else f"Col {i}" for i, x in enumerate(raw_h)]
    except Exception:
        headers = [f"Column_{i}" for i in range(dims["cols"])]

    return {
        "workbook_id": wb_id,
        "meta": {
            "rows": max(1, dims["rows"] - offset - 1),
            "cols": dims["cols"],
            "offset": offset,
            "headers": headers,
            "title": ""
        }
    }

@app.get("/api/workbook/{workbook_id}/rows")
async def get_rows(workbook_id: str, start: int, count: int, offset: int):
    files = list(UPLOAD_DIR.glob(f"{workbook_id}.*"))
    if not files: raise HTTPException(status_code=404, detail="Session expired.")
    
    file_path = files[0]
    suffix = file_path.suffix.lower()
    adjusted_start = offset + 1 + start
    
    try:
        if suffix == '.csv':
            df = processor.parse_csv_with_encoding_fallback(file_path, nrows=count, skiprows=adjusted_start)
        else:
            df = pd.read_excel(file_path, engine='openpyxl', skiprows=adjusted_start, nrows=count, header=None)
        
        matrix = df.fillna("").values.tolist()
        
        # CRITICAL RUNTIME FIX 2: Check backend state cache memory to override loaded baseline columns
        deltas = SESSION_DELTAS.get(workbook_id, {}).get("data", {})
        for r_idx in range(len(matrix)):
            actual_data_row = start + r_idx
            for c_idx in range(len(matrix[r_idx])):
                delta_key = f"{actual_data_row}-{c_idx}"
                if delta_key in deltas:
                    matrix[r_idx][c_idx] = deltas[delta_key]
                    
        return {"rows": matrix, "start": start}
    except Exception:
        return {"rows": [], "start": start}

@app.post("/api/mutate/{workbook_id}")
async def mutate_session(workbook_id: str, payload: dict):
    if workbook_id not in SESSION_DELTAS:
        raise HTTPException(status_code=404, detail="Session not found.")
    if "data" in payload: SESSION_DELTAS[workbook_id]["data"].update(payload["data"])
    if "headers" in payload: SESSION_DELTAS[workbook_id]["headers"].update(payload["headers"])
    if "styles" in payload: SESSION_DELTAS[workbook_id]["styles"].update(payload["styles"])
    if "added_rows_count" in payload: SESSION_DELTAS[workbook_id]["added_rows_count"] = payload["added_rows_count"]
    if "added_cols_count" in payload: SESSION_DELTAS[workbook_id]["added_cols_count"] = payload["added_cols_count"]
    if "title" in payload: SESSION_DELTAS[workbook_id]["title"] = payload["title"]
    
    if "headers_array" in payload:
        for i, val in enumerate(payload["headers_array"]):
            SESSION_DELTAS[workbook_id]["headers"][str(i)] = val
            
    return {"status": "synced"}

@app.post("/api/export/{workbook_id}")
async def export_workbook(workbook_id: str, payload: dict = None):
    files = list(UPLOAD_DIR.glob(f"{workbook_id}.*"))
    if not files: raise HTTPException(status_code=404, detail="File missing.")
    
    file_path = files[0]
    suffix = file_path.suffix.lower()
    deltas = SESSION_DELTAS.get(workbook_id, {"data": {}, "headers": {}, "styles": {}, "added_rows_count": 0, "added_cols_count": 0, "base_offset": 0, "title": ""})
    offset = deltas.get("base_offset", 0)

    theme = (payload or {}).get("theme", "emerald")
    title = (payload or {}).get("title", deltas.get("title", ""))
    
    signatory = "SOFTWARE: EXCEL DESIGNER; VERSION: 1.00; AUTHOR: ANKUR DOWARAH; URL: ank-ldb.blogspot.com"

    excel_theme_fills = {
        "progress": {"header": "F8FAFC", "even": "FFFFFF", "odd": "FDFBF7"},
        "teal-dashboard": {"header": "005F56", "even": "FFFFFF", "odd": "FAFAFA"},
        "financial": {"header": "104F60", "even": "E8EEF3", "odd": "FFFFFF"},
        "emerald": {"header": "107C41", "even": "FFFFFF", "odd": "F8FAFC"}
    }
    t_set = excel_theme_fills.get(theme, excel_theme_fills["emerald"])

    try:
        if suffix == '.csv':
            df = processor.parse_csv_with_encoding_fallback(file_path, header=None)
            wb = openpyxl.Workbook()
            ws = wb.active
            for r_idx, row in enumerate(df.values.tolist()):
                for c_idx, val in enumerate(row):
                    ws.cell(row=r_idx+1, column=c_idx+1, value=val)
        else:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
        # Structure modifications sizing expansions
        added_rows = deltas.get("added_rows_count", 0)
        added_cols = deltas.get("added_cols_count", 0)
        
        for c_str, h_val in deltas.get("headers", {}).items():
            ws.cell(row=offset + 1, column=int(c_str) + 1, value=h_val)

        for key, value in deltas.get("data", {}).items():
            r, c = map(int, key.split("-"))
            ws.cell(row=offset + 2 + r, column=c + 1, value=value)
            
        max_r = ws.max_row + added_rows
        max_c = ws.max_column + added_cols
        
        # Style Headers
        for c in range(1, max_c + 1):
            cell = ws.cell(row=offset + 1, column=c)
            cell.fill = PatternFill(start_color=t_set["header"], end_color=t_set["header"], fill_type="solid")
            if theme in ["teal-dashboard", "financial", "emerald"]:
                cell.font = Font(color="FFFFFF", bold=True)
        
        # Style Uniform Data Body Rows
        for r in range(offset + 2, max_r + 1):
            row_color = t_set["even"] if ((r - offset) % 2 == 0) else t_set["odd"]
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                
                style_key = f"{r-offset-1}-{c-1}"
                cell_styles = deltas.get("styles", {}).get(style_key, {})
                if cell_styles.get("bold"):
                    cell.font = Font(bold=True)

        ws.cell(row=max_r + 2, column=1, value=f"Title Context: {title}" if title else "")
        ws.cell(row=max_r + 3, column=1, value=signatory)

        timestamp = datetime.now().strftime("%d%m%Y%H%M")
        filename = f"AnkLDB_excel-Designer_{timestamp}.xlsx"
            
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        return StreamingResponse(
            out, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logger.error(f"Export Error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))