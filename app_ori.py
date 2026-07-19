import os
import uuid
import logging
import subprocess
import openpyxl
import io
import shutil
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
import pandas as pd

app = FastAPI()
# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# grid 
class GridService:
    """Handles high-performance chunck reading for the virtualized frontend."""
    @staticmethod
    def get_chunck(file_path: Path, start: int, count: int):
        wb = openpyxl.load_workbook(file_path, read_only=True, date_only=True). HTML
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=start + 1, max_row=start + count, values_only=True):
            rows.append(list(row))
        return rows
# Config
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
MAX_SIZE = 10 * 1024 * 1024  # 10MB limit (adjust as needed)
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

class AnkDO_Processor:
    def validate_file(self, filename: str):
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail="Unsupported file type.")
        return suffix

    def get_metadata(self, file_path: str, suffix: str):
        """Extracts dimensions without loading the full data (Ankur's method)."""
        if suffix == '.csv':
                return {"rows": sum(1 for _ in open(file_path)), "cols":0}
        else:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                return {"rows": ws.max_row, "cols": ws.max_column}

    def run_heavy_ocr(self, image_path: str):
        try:
            result = subprocess.run(
                ['python', 'utils/ocr_worker.py', image_path],
                capture_output=True, text=True, check=True, timeout=300
            )
            return result.stdout
        except subprocess.CalledProcessError as e:
            logger.error(f"OCR failed: {e.stderr}")
            return None

ank_designer = AnkDO_Processor()
@app.get ("/")
async def serve_index():
    return FileResponse("templates/index.html")
@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    suffix = ank_designer.validate_file(file.filename)
    wb_id = str(uuid.uuid4())
    temp_path = UPLOAD_DIR / f"{wb_id}{suffix}"
    
    # Save file permanently (DO NOT delete in finally block)
    with open(temp_path, "wb") as f:
        f.write(await file.read())
    
    metadata = ank_designer.get_metadata(temp_path, suffix)
    return {
        "status": "success", 
        "workbook_id": wb_id, 
        "meta": metadata
    }
@app.post("/api/export")
async def export_data(data: dict):
    # Flatten the JSON input back into a DataFrame
    df = pd.DataFrame(data['data'])
    
    # Create an in-memory Excel file
    output = io.BytesIO()
    df.to_excel(output, index=False, header=False)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=workbook_export.xlsx"}
    )
@app.get("/api/workbook/{workbook_id}/rows")
async def get_rows(workbook_id: str, start: int, count: int):
    # Find file by ID (any extension)
    files = list(UPLOAD_DIR.glob(f"{workbook_id}.*"))
    if not files:
        raise HTTPException(status_code=404, detail="Workbook not found.")
    
    file_path = files[0]
    try:
        # Load slice
        df = pd.read_excel(file_path, engine='openpyxl', skiprows=range(1, start + 1), nrows=count, header=None)
        return {"rows": df.values.tolist(), "start": start}
    except Exception as e:
        logger.error(f"Error fetching rows: {e}")
        raise HTTPException(status_code=500, detail="Error fetching data chunk.")
    # Construct the path based on your uuid naming convention
    file_path = UPLOAD_DIR / f"{workbook_id}.xlsx" # Or detect extension
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Workbook not found.")
        
    try:
        df = pd.read_excel(
            file_path, 
            engine='openpyxl', 
            skiprows=range(1, start + 1), 
            nrows=count,
            header=None # Assuming headers were already fetched in metadata
        )
        return {"rows": df.values.tolist()}
    except Exception as e:
        logger.error(f"Error fetching rows: {e}")
        raise HTTPException(status_code=500, detail="Error fetching data chunk.")