import os
import uuid
import logging
import io
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import openpyxl

app = FastAPI()

# Enable CORS so the browser doesn't block local frontend requests
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Config
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

class AnkDO_Processor:
    def validate_file(self, filename: str):
        suffix = Path(filename).suffix.lower()
        if suffix not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail="Unsupported file type.")
        return suffix

    def get_metadata(self, file_path: Path, suffix: str):
        """Extracts accurate dimensions for both CSV and Excel files."""
        try:
            if suffix == '.csv':
                # Read first row to determine columns, count rows efficiently
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    first_line = f.readline()
                    cols = len(first_line.split(',')) if first_line else 0
                    f.seek(0)
                    rows = sum(1 for _ in f)
                return {"rows": rows, "cols": cols if cols > 0 else 10}
            else:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                # Fallback to dataframe measurement if max_row is inaccurate
                rows = ws.max_row if ws.max_row else 1000
                cols = ws.max_column if ws.max_column else 26
                return {"rows": rows, "cols": cols}
        except Exception as e:
            logger.error(f"Metadata extraction failed: {e}")
            return {"rows": 100, "cols": 10}

ank_designer = AnkDO_Processor()

# Serve static files from a 'templates' folder if needed
@app.get("/")
async def serve_index():
    if os.path.exists("templates/index.html"):
        return FileResponse("templates/index.html")
    elif os.path.exists("index.html"):
        return FileResponse("index.html")
    raise HTTPException(status_code=404, detail="index.html template layout not found.")

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    suffix = ank_designer.validate_file(file.filename)
    wb_id = str(uuid.uuid4())
    temp_path = UPLOAD_DIR / f"{wb_id}{suffix}"
    
    try:
        with open(temp_path, "wb") as f:
            f.write(await file.read())
        
        metadata = ank_designer.get_metadata(temp_path, suffix)
        return {
            "status": "success", 
            "workbook_id": wb_id, 
            "meta": metadata
        }
    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"File save failure: {str(e)}")

@app.post("/api/export")
async def export_data(data: dict):
    try:
        df = pd.DataFrame(data['data'])
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=workbook_export.xlsx"}
        )
    except Exception as e:
        logger.error(f"Export error: {e}")
        raise HTTPException(status_code=500, detail="Failed to compile export sheet.")

@app.get("/api/workbook/{workbook_id}/rows")
async def get_rows(workbook_id: str, start: int, count: int):
    # Find matching file by ID regardless of extension
    files = list(UPLOAD_DIR.glob(f"{workbook_id}.*"))
    if not files:
        raise HTTPException(status_code=404, detail="Workbook file reference not found.")
    
    file_path = files[0]
    suffix = file_path.suffix.lower()
    
    try:
        if suffix == '.csv':
            # Fast-chunk reading for CSV
            df = pd.read_csv(file_path, skiprows=start, nrows=count, header=None)
        else:
            # Fast-chunk reading for Excel sheets
            df = pd.read_excel(file_path, engine='openpyxl', skiprows=start, nrows=count, header=None)
            
        # Replace NaN values with empty strings so JSON remains valid
        df = df.fillna("")
        return {"rows": df.values.tolist(), "start": start}
    except Exception as e:
        logger.error(f"Error fetching rows: {e}")
        # Return empty list structure if chunk window falls outside real bounds gracefully
        return {"rows": [], "start": start}